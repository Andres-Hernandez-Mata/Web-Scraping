import os, requests, re, json
from openpyxl.worksheet.table import Table, TableStyleInfo
from geopy.geocoders import Nominatim

try:
	from bs4 import BeautifulSoup as bs
except:
	os.system("pip install beautifulsoup4")
	os.system("clear")
try:
	from openpyxl import load_workbook, Workbook
except:
	os.system("pip install openpyxl")
	os.system("clear")


def soup_re(urls_list,search_query):
	app = Nominatim(user_agent="tutorial")
	# Archivo excel
	workbook = Workbook()
	# Nombre del excel
	file_name = "Web_Scraping.xlsx"
	# Se activa la hoja para poder trabajar en ella
	excel_page = workbook.active
	# Nombre del hoja
	excel_page.title = search_query
	# Columnas del excel
	excel_page.append(('Url','Email','Phone Number','Social Network','News'))
	# Guardar el nombre de la noticia
	news = ""
	
	# Se crean las expresiones regulares
	email_regex = re.compile(r"[A-z0-9]+[\._]?[A-z0-9]+@[A-z]+.com")
	phone_regex = re.compile(r"\+?\d?\d?\s?\(?\d{2}\)?\s\d{4}-?\s?\d{4}")
	social_regex = re.compile(r"@\w*"+search_query.lower()+"\w*")

	for url in urls_list:
		page = requests.get(url)

		# Encontrar todas las coincidencias
		match_email = email_regex.findall(page.text)
		match_phone = phone_regex.findall(page.text)
		match_social = social_regex.findall(page.text)

		soup = bs(page.content, 'html.parser')

		# Obtenemos el origen del atleta/equipo
		th = soup.select("th")
		for tag_origen in th:
			if "Nacimiento" in tag_origen.getText():
				nacimiento = soup.find('a', class_='mw-redirect')
				location = app.geocode(nacimiento.getText()).raw

		# Obtenemos los encabezamos
		h1 = soup.select("h1")
		for tag_news in h1:
			if search_query in tag_news.getText():
				# Obtenemos el nombre de la noticia para guardarlo
				news = tag_news.getText()

		# Practicamente lo que se requiere es crear una cadena mediante una lista
		email = ','.join(match_email)
		phone = ','.join(match_phone)
		social = ','.join(match_social)

		# Manejo de estructura de datos	
		# Se crea una tupla para almancer la informacion
		tupla_informacion = tuple()
		tupla_informacion = (url, email, phone, social, news)
		# Se crea una lista para que con un ciclo recorrer una fila y agregar la informacion en columnas correspondientes
		info_list = list() 
		info_list.append(tupla_informacion)	
		for info in info_list:		
			excel_page.append(info)
	
	# Creamos una tabla como estilo en el excel
	tab = Table(displayName="Tabla", ref="A1:E5")
	style = TableStyleInfo(name=search_query, showFirstColumn=True, showLastColumn=True, showRowStripes=True, showColumnStripes=True)
	tab.tableStyleInfo = style
	excel_page.add_table(tab)

	try:
		# Obtenemos la respuesta para la localizacion 
		json_str = json.dumps(location)
		resp = json.loads(json_str)
	except ValueError:
		pass

	# Creamos la url con la informacion indicada
	url = "https://api.openweathermap.org/data/2.5/onecall?lat="+resp['lat']+"&lon="+resp['lon']+"&exclude=minutely,hourly,daily&units=metric&appid="
	# Se realiza la peticion
	res = requests.get(url)
	# Se obtiene el contenido
	clima = json.loads(res.content)
	# Crear una nueva hoja
	excel_page = workbook.create_sheet("Clima") 
	# Obtenemos la informacion del json
	temperatura = clima["current"]["temp"]
	# Agregar la informacion de la temperatura
	excel_page.append(('Temperatura',temperatura))
	# Guardar el archivo con la informacion		
	workbook.save(filename=file_name)
	print("Listo puedes consultar el archivo "+file_name+" con la informacion obtenida")

# Se descargan las imagenes
def soup_img(urls_list,search_query):
	i = 0
	for urls in urls_list:
		page = requests.get(urls)
		soup = bs(page.content, 'html.parser')
		tags = soup.find_all("img")
		for tag in tags:
			link_img = tag.get("src")
			try:
				img = requests.get(link_img).content				
				with open("img/"+search_query+"/"+str(i)+".jpg", "wb") as file:
					file.write(img)
			except OSError:
				continue
			i += 1
	print("Imagenes descargadas de "+search_query)

